from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base_parser import ParsedDocument


ORG_TO_SERVICE = {
    "A": "US Army",
    "N": "US Navy",
    "F": "US Air Force",
    "M": "US Marine Corps",
    "D": "Defense-Wide",
    "S": "US Space Force",
}


@dataclass(frozen=True)
class DisplayWorkbookConfig:
    document_type: str
    document_id: str
    title: str
    appropriation_type: str
    sheet_candidates: tuple[str, ...]
    source_filename: str


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "row"


def _coerce_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _thousands_to_millions(value: Any) -> float | None:
    number = _coerce_number(value)
    if number is None:
        return None
    return round(number / 1000, 3)


def _quantity(value: Any) -> float | None:
    number = _coerce_number(value)
    if number is None:
        return None
    return round(number, 3)


def _first(raw: dict[str, Any], names: list[str]) -> Any:
    for name in names:
        if name in raw:
            return raw[name]
    return None


def _infer_service(organization: Any, account_title: Any) -> str | None:
    org = str(organization or "").strip().upper()
    if org in ORG_TO_SERVICE:
        return ORG_TO_SERVICE[org]

    title = str(account_title or "").lower()
    if "space force" in title:
        return "US Space Force"
    if "air force" in title:
        return "US Air Force"
    if "navy" in title:
        return "US Navy"
    if "marine corps" in title:
        return "US Marine Corps"
    if "army" in title:
        return "US Army"
    if "defense-wide" in title or "defense wide" in title:
        return "Defense-Wide"
    return None


def _is_included_in_toa(value: Any, document_type: str) -> bool:
    text = str(value or "").strip().lower()
    if document_type == "p1":
        return text == "add"
    return text == "y"


def _needs_review(raw: dict[str, Any], amount: float | None, line_name: str | None) -> bool:
    if not line_name:
        return True
    if amount is None:
        return True
    classification = str(raw.get("Classification") or "").upper()
    return classification not in {"U", "C", "S", "TS", ""}


class DisplayWorkbookParser:
    def __init__(self, config: DisplayWorkbookConfig):
        self.config = config

    def parse(self, workbook_path: str) -> ParsedDocument:
        path = Path(workbook_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = next((name for name in self.config.sheet_candidates if name in workbook.sheetnames), workbook.sheetnames[0])
        sheet = workbook[sheet_name]

        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value).strip() if value is not None else "" for value in rows[1]]
        total_row = self._row_to_raw(headers, rows[0])
        data_rows = rows[2:]

        line_items: list[dict[str, Any]] = []
        for row_index, row in enumerate(data_rows, start=3):
            raw = self._row_to_raw(headers, row)
            if not self._is_data_row(raw):
                continue
            line_items.append(self._line_item(raw, row_index, sheet_name))

        included_total = round(
            sum(item["amount_millions"] or 0 for item in line_items if item["include_in_toa"]),
            3,
        )
        displayed_total = _thousands_to_millions(
            _first(total_row, ["FY 2027 Total", "FY 2027 Total Amount"]),
        )

        metadata = {
            "id": self.config.document_id,
            "title": self.config.title,
            "fiscal_year": 2027,
            "document_type": self.config.document_type,
            "service_or_component": None,
            "appropriation_type": self.config.appropriation_type,
            "source_filename": self.config.source_filename,
            "source_url": None,
            "sheet_name": sheet_name,
            "row_count": len(line_items),
            "included_in_toa_total_millions": included_total,
            "displayed_total_millions": displayed_total,
            "parser_confidence": 0.96,
        }
        return ParsedDocument(metadata=metadata, line_items=line_items)

    def _row_to_raw(self, headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
        return {
            header: value
            for header, value in zip(headers, row)
            if header and value is not None and value != ""
        }

    def _is_data_row(self, raw: dict[str, Any]) -> bool:
        return bool(raw.get("Account") and (raw.get("Line Number") or raw.get("PE/BLI") or raw.get("Budget Line Item")))

    def _line_item(self, raw: dict[str, Any], row_index: int, sheet_name: str) -> dict[str, Any]:
        line_name = _first(
            raw,
            [
                "Program Element/Budget Line Item (BLI) Title",
                "Budget Line Item (BLI) Title",
                "SAG/Budget Line Item (BLI) Title",
            ],
        )
        program_element = _first(raw, ["PE/BLI", "Budget Line Item", "SAG/BLI"])
        bsa = _first(raw, ["BSA", "AG/BSA"])
        bsa_name = _first(raw, ["Budget SubActivity (BSA) Title", "AG/Budget SubActivity (BSA) Title"])
        include_raw = _first(raw, ["Include In TOA", "Add/Non-Add"])
        amount = _thousands_to_millions(_first(raw, ["FY 2027 Total", "FY 2027 Total Amount"]))
        quantity = _quantity(_first(raw, ["FY 2027 Total Quantity"]))
        account = str(raw.get("Account") or "")
        line_number = str(raw.get("Line Number") or "")
        item_id = _slug(
            "-".join(
                [
                    self.config.document_id,
                    account,
                    line_number,
                    str(program_element or ""),
                    str(raw.get("Cost Type") or ""),
                    f"r{row_index}",
                ],
            ),
        )

        service = _infer_service(raw.get("Organization"), raw.get("Account Title"))
        fy2025_total = _thousands_to_millions(_first(raw, ["FY 2025 Total", "FY 2025 Total Amount"]))
        fy2026_total = _thousands_to_millions(_first(raw, ["FY 2026 Total", "FY 2026 Total Amount"]))

        return {
            "id": item_id,
            "document_id": self.config.document_id,
            "document_type": self.config.document_type,
            "program_id": None,
            "mission_area_id": None,
            "service_or_component": service,
            "appropriation_type": self.config.appropriation_type,
            "appropriation_account": account,
            "account_title": raw.get("Account Title"),
            "organization": raw.get("Organization"),
            "budget_activity": str(raw.get("Budget Activity") or ""),
            "budget_activity_name": raw.get("Budget Activity Title"),
            "budget_subactivity": str(bsa or ""),
            "budget_subactivity_name": bsa_name,
            "program_element": str(program_element or ""),
            "line_number": line_number,
            "line_item_name": str(line_name or ""),
            "project_number": None,
            "project_name": None,
            "cost_type": raw.get("Cost Type"),
            "cost_type_title": raw.get("Cost Type Title"),
            "add_non_add": raw.get("Add/Non-Add"),
            "include_in_toa": _is_included_in_toa(include_raw, self.config.document_type),
            "include_in_toa_label": str(include_raw or ""),
            "classification": raw.get("Classification"),
            "fiscal_year": 2027,
            "prior_year_actual": fy2025_total,
            "current_year_enacted": fy2026_total,
            "budget_year_request": amount,
            "amount_millions": amount,
            "quantity": quantity,
            "fy2025_actual_amount_millions": _thousands_to_millions(
                _first(raw, ["FY 2025 Actuals", "FY 2025 Actuals Amount"]),
            ),
            "fy2025_reconciliation_amount_millions": _thousands_to_millions(
                _first(raw, ["FY 2025 Reconciliation", "FY 2025 Reconciliation Amount"]),
            ),
            "fy2025_total_amount_millions": fy2025_total,
            "fy2026_discretionary_enacted_amount_millions": _thousands_to_millions(
                _first(raw, ["FY 2026 Discretionary Enacted", "FY 2026 Discretionary Enacted Amount"]),
            ),
            "fy2026_mandatory_amount_millions": _thousands_to_millions(
                _first(raw, ["FY 2026 PL 119-21 Spend Plan", "FY 2026 PL 119-21 Spend Plan Amount"]),
            ),
            "fy2026_total_amount_millions": fy2026_total,
            "fy2027_discretionary_request_amount_millions": _thousands_to_millions(
                _first(raw, ["FY 2027 Discretionary Request", "FY 2027 Discretionary Request Amount"]),
            ),
            "fy2027_mandatory_request_amount_millions": _thousands_to_millions(
                _first(raw, ["FY 2027 Mandatory Request", "FY 2027 Mandatory Amount"]),
            ),
            "fy2027_total_amount_millions": amount,
            "fy2025_actual_quantity": _quantity(_first(raw, ["FY 2025 Actuals Quantity"])),
            "fy2025_reconciliation_quantity": _quantity(_first(raw, ["FY 2025 Reconcilation Quantity"])),
            "fy2025_total_quantity": _quantity(_first(raw, ["FY 2025 Total Quantity"])),
            "fy2026_discretionary_enacted_quantity": _quantity(_first(raw, ["FY 2026 Discretionary Enacted Quantity"])),
            "fy2026_mandatory_quantity": _quantity(_first(raw, ["FY 2026  PL 119-21 Spend Plan Quantity"])),
            "fy2026_total_quantity": _quantity(_first(raw, ["FY 2026 Total Quantity"])),
            "fy2027_discretionary_request_quantity": _quantity(_first(raw, ["FY 2027 Discretionary Request Quantity"])),
            "fy2027_mandatory_request_quantity": _quantity(_first(raw, ["FY 2027 Mandatory Quantity"])),
            "fy2027_total_quantity": quantity,
            "source_page": f"{sheet_name} row {row_index}",
            "raw_text": json.dumps(raw, default=str, sort_keys=True),
            "raw_row": raw,
            "confidence_score": 0.95 if amount is not None and line_name else 0.7,
            "needs_review": _needs_review(raw, amount, str(line_name or "")),
        }
